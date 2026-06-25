import logging
from typing import Optional
from datetime import date, datetime
from database import get_connection

logger = logging.getLogger(__name__)


class ReportController:
    def laporan_harian(self, tanggal: Optional[str] = None) -> dict:
        try:
            if not tanggal:
                tanggal = date.today().isoformat()

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    """SELECT a.id, u.nama AS mahasiswa, u.username,
                              a.jam_masuk, a.jam_pulang, a.status, a.keterangan
                       FROM absensi a
                       JOIN users u ON a.user_id = u.id
                       WHERE a.tanggal = ?
                       ORDER BY u.nama ASC""",
                    (tanggal,)
                )
                rows = [dict(r) for r in cursor.fetchall()]

                cursor.execute(
                    """SELECT COUNT(*) AS total,
                              SUM(CASE WHEN status = 'hadir' THEN 1 ELSE 0 END) AS hadir,
                              SUM(CASE WHEN status = 'izin' THEN 1 ELSE 0 END) AS izin,
                              SUM(CASE WHEN status = 'sakit' THEN 1 ELSE 0 END) AS sakit,
                              SUM(CASE WHEN status = 'alpha' THEN 1 ELSE 0 END) AS alpha
                       FROM absensi WHERE tanggal = ?""",
                    (tanggal,)
                )
                stat = dict(cursor.fetchone())

            return {'success': True, 'tanggal': tanggal, 'data': rows, 'statistik': stat}
        except Exception as e:
            logger.error("Gagal buat laporan harian: %s", e, exc_info=True)
            return {'success': False, 'message': 'Terjadi kesalahan sistem'}

    def laporan_bulanan(self, bulan: Optional[int] = None,
                        tahun: Optional[int] = None) -> dict:
        try:
            now = date.today()
            if not bulan:
                bulan = now.month
            if not tahun:
                tahun = now.year

            bulan_str = f"{tahun}-{bulan:02d}"

            with get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    """SELECT u.id, u.nama, u.username,
                              COUNT(a.id) AS total_hari,
                              SUM(CASE WHEN a.status = 'hadir' THEN 1 ELSE 0 END) AS hadir,
                              SUM(CASE WHEN a.status = 'izin' THEN 1 ELSE 0 END) AS izin,
                              SUM(CASE WHEN a.status = 'sakit' THEN 1 ELSE 0 END) AS sakit,
                              SUM(CASE WHEN a.status = 'alpha' THEN 1 ELSE 0 END) AS alpha
                       FROM users u
                       LEFT JOIN absensi a ON u.id = a.user_id
                           AND strftime('%Y-%m', a.tanggal) = ?
                       WHERE u.role = 'mahasiswa'
                       GROUP BY u.id, u.nama, u.username
                       ORDER BY u.nama ASC""",
                    (bulan_str,)
                )
                rows = [dict(r) for r in cursor.fetchall()]

            return {'success': True, 'bulan': bulan, 'tahun': tahun, 'data': rows}
        except Exception as e:
            logger.error("Gagal buat laporan bulanan: %s", e, exc_info=True)
            return {'success': False, 'message': 'Terjadi kesalahan sistem'}

    def export_excel(self, data: list[dict], filename: Optional[str] = None) -> dict:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError as e:
            logger.error("openpyxl tidak terinstall: %s", e)
            return {'success': False,
                    'message': 'Library openpyxl belum terpasang. Jalankan: pip install openpyxl'}

        try:
            if not data:
                return {'success': False, 'message': 'Tidak ada data untuk diexport'}

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Laporan Absensi'

            headers = list(data[0].keys())
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                                      fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx,
                                   value=row_data.get(header, ''))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

            for col_idx, header in enumerate(headers, 1):
                max_len = len(str(header))
                for row_idx in range(2, len(data) + 2):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = max_len + 4

            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'laporan_absensi_{timestamp}.xlsx'

            wb.save(filename)
            logger.info("Excel exported: %s (%d rows)", filename, len(data))
            return {'success': True, 'message': f'File berhasil disimpan: {filename}',
                    'filename': filename}
        except Exception as e:
            logger.error("Gagal export Excel: %s", e, exc_info=True)
            return {'success': False, 'message': f'Gagal export Excel: {str(e)}'}

    def export_pdf(self, data: list[dict], filename: Optional[str] = None,
                   title: str = 'Laporan Absensi') -> dict:
        try:
            from fpdf import FPDF
        except ImportError as e:
            logger.error("fpdf2 tidak terinstall: %s", e)
            return {'success': False,
                    'message': 'Library fpdf2 belum terpasang. Jalankan: pip install fpdf2'}

        try:
            if not data:
                return {'success': False, 'message': 'Tidak ada data untuk diexport'}

            pdf = FPDF(orientation='L', unit='mm', format='A4')
            pdf.add_page()
            pdf.set_font('Helvetica', 'B', 16)
            pdf.cell(0, 10, title, new_x='LMARGIN', new_y='NEXT', align='C')
            pdf.ln(5)

            headers = list(data[0].keys())
            col_width = (270 - 10) / len(headers) if headers else 50

            pdf.set_font('Helvetica', 'B', 9)
            pdf.set_fill_color(68, 114, 196)
            pdf.set_text_color(255, 255, 255)
            for header in headers:
                pdf.cell(col_width, 8, header.replace('_', ' ').title(),
                         border=1, align='C', fill=True)
            pdf.ln()

            pdf.set_font('Helvetica', '', 8)
            pdf.set_text_color(0, 0, 0)
            fill = False
            for row_data in data:
                pdf.set_fill_color(220, 220, 220) if fill else pdf.set_fill_color(255, 255, 255)
                for header in headers:
                    pdf.cell(col_width, 7, str(row_data.get(header, '')),
                             border=1, align='C', fill=True)
                pdf.ln()
                fill = not fill

            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f'laporan_absensi_{timestamp}.pdf'

            pdf.output(filename)
            logger.info("PDF exported: %s (%d rows)", filename, len(data))
            return {'success': True, 'message': f'File berhasil disimpan: {filename}',
                    'filename': filename}
        except Exception as e:
            logger.error("Gagal export PDF: %s", e, exc_info=True)
            return {'success': False, 'message': f'Gagal export PDF: {str(e)}'}
